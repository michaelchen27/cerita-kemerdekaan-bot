from bs4 import BeautifulSoup
import selenium.webdriver as webdriver
import openpyxl as xl
import os

edited_excel_path = "D:\\IG_Project\\checked_form.xlsx"

if os.path.exists(edited_excel_path):
    os.remove(edited_excel_path)

wb = xl.load_workbook('sorted_form.xlsx')
sheet = wb['Form Responses 1']


counter = 0
for row in range(2, sheet.max_row + 1):

    driver = webdriver.Chrome(executable_path="D:\\Downloads\\chromedriver.exe")

    url = sheet.cell(row, 12).value
    driver.get(url)
    soup = BeautifulSoup(driver.page_source, features="lxml")

    profile_url = sheet.cell(row, 10).value
    driver.get(profile_url)
    soup_profile = BeautifulSoup(driver.page_source, features="lxml")

    username = soup_profile.find("title").text
    username = username.split('• ')[0]
    counter += 1
    print(f'{counter}.{username}\n----------------------------------------')


# Cek Tag @djikp
    tagExist = False
    tagCounter = 0

    for tag in soup.find_all('a', class_='notranslate'):
        if tag.text == "@djikp":
            tagExist = True
        tagCounter += 1

    if tagExist and tagCounter >= 6:
        sheet.cell(row, 18).value = "ADA"

    else:
        sheet.cell(row, 18).value = "TIDAK ADA"

    print(f'People tagged\t: {tagCounter}')

# Cek hashtag
    i = 0
    for x in soup.find_all('a', {'class': 'xil3i'}):
        if x.text.lower() == "#ceritakemerdekaan" or x.text.lower() == "#lombapodcastkominfo":
            i += 1

    if i >= 2:
        sheet.cell(row, 20).value = "ADA"
        print(f'Hashtag\t\t\t: Lengkap')

    else:
        sheet.cell(row, 20).value = "TIDAK ADA"
        print('Hashtag\t\t\t: Tidak lengkap')

# Cek Durasi Video
    video = soup.find('video')
    if video:
        video_url = video['src']
        print(f'Video URL\t\t: {video_url}')
    else:
        print('Video URL\t\t: null')

    wb.save("checked_form.xlsx")

    print('DONE CHECKING!\n')