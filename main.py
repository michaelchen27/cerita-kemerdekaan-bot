import openpyxl as xl
import webbrowser
import time

wb = xl.load_workbook('sorted_form.xlsx')
sheet = wb['Form Responses 1']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 10)
    print(cell)